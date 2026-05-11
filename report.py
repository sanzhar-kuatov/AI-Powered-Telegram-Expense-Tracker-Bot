import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import (get_expenses_by_period, get_category_summary,
                get_income_by_period, get_income_summary)


# ─── Style Helpers ────────────────────────────────────────────

def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def _header_style(cell, bg_color: str = "2F4F7F"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin_border()

def _cell_style(cell, bold: bool = False, align: str = "left", bg: str = None):
    cell.font = Font(bold=bold, name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def _set_col_width(sheet, col: int, width: float):
    sheet.column_dimensions[get_column_letter(col)].width = width


# ─── Table Builders ───────────────────────────────────────────

def _build_summary_table(sheet, summary_rows: list,
                          total_income: float, start_row: int) -> int:
    title = sheet.cell(row=start_row, column=1, value="📊 Spending by Category")
    title.font = Font(bold=True, name="Arial", size=13)
    start_row += 1

    for col, header in enumerate(["Category", "Total Spent (RM)", "% of Total"], 1):
        _header_style(sheet.cell(row=start_row, column=col, value=header))
    start_row += 1

    data_start = start_row
    total_expense_row = data_start + len(summary_rows)

    for i, (category, total) in enumerate(summary_rows):
        r = data_start + i
        _cell_style(sheet.cell(row=r, column=1, value=category))
        amt = sheet.cell(row=r, column=2, value=round(float(total), 2))
        _cell_style(amt, align="right")
        amt.number_format = '#,##0.00'
        pct = sheet.cell(row=r, column=3, value=f"=B{r}/B{total_expense_row}*100")
        _cell_style(pct, align="right")
        pct.number_format = '0.0"%"'

    # Total expenses row
    _cell_style(sheet.cell(row=total_expense_row, column=1,
                            value="TOTAL EXPENSES"), bold=True)
    te = sheet.cell(row=total_expense_row, column=2,
                    value=f"=SUM(B{data_start}:B{total_expense_row-1})")
    _cell_style(te, bold=True, align="right")
    te.number_format = '#,##0.00'
    sheet.cell(row=total_expense_row, column=3, value="100.0%")

    # Total income row
    income_row = total_expense_row + 1
    _cell_style(sheet.cell(row=income_row, column=1,
                            value="TOTAL INCOME"), bold=True, bg="E8F5E9")
    inc = sheet.cell(row=income_row, column=2, value=round(total_income, 2))
    _cell_style(inc, bold=True, align="right", bg="E8F5E9")
    inc.number_format = '#,##0.00'

    # Net balance row
    net_row = income_row + 1
    _cell_style(sheet.cell(row=net_row, column=1,
                            value="NET BALANCE"), bold=True, bg="FFF9C4")
    net = sheet.cell(row=net_row, column=2,
                     value=f"=B{income_row}-B{total_expense_row}")
    _cell_style(net, bold=True, align="right", bg="FFF9C4")
    net.number_format = '#,##0.00'

    return net_row + 2


def _build_income_table(sheet, income_rows: list, start_row: int) -> int:
    title = sheet.cell(row=start_row, column=1, value="💵 Income Details")
    title.font = Font(bold=True, name="Arial", size=13)
    start_row += 1

    for col, header in enumerate(["Date", "Description", "Amount (RM)"], 1):
        _header_style(sheet.cell(row=start_row, column=col, value=header),
                      bg_color="1F6B3E")
    start_row += 1

    for i, (description, amount, income_date) in enumerate(income_rows):
        r = start_row + i
        bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"
        date_cell = sheet.cell(row=r, column=1,
                               value=income_date.strftime("%d/%m/%Y")
                               if hasattr(income_date, "strftime") else str(income_date))
        _cell_style(date_cell, align="center", bg=bg)
        _cell_style(sheet.cell(row=r, column=2, value=description), bg=bg)
        amt = sheet.cell(row=r, column=3, value=round(float(amount), 2))
        _cell_style(amt, align="right", bg=bg)
        amt.number_format = '#,##0.00'

    return start_row + len(income_rows) + 2


def _build_expenses_table(sheet, expense_rows: list, start_row: int) -> int:
    title = sheet.cell(row=start_row, column=1, value="🧾 Expense Details")
    title.font = Font(bold=True, name="Arial", size=13)
    start_row += 1

    for col, header in enumerate(["Date", "Description", "Category", "Amount (RM)"], 1):
        _header_style(sheet.cell(row=start_row, column=col, value=header),
                      bg_color="7B3F00")
    start_row += 1

    for i, (description, amount, category_name, expense_date) in enumerate(expense_rows):
        r = start_row + i
        bg = "F5F5F5" if i % 2 == 0 else "FFFFFF"
        date_cell = sheet.cell(row=r, column=1,
                               value=expense_date.strftime("%d/%m/%Y")
                               if hasattr(expense_date, "strftime") else str(expense_date))
        _cell_style(date_cell, align="center", bg=bg)
        _cell_style(sheet.cell(row=r, column=2, value=description), bg=bg)
        _cell_style(sheet.cell(row=r, column=3, value=category_name),
                    align="center", bg=bg)
        amt = sheet.cell(row=r, column=4, value=round(float(amount), 2))
        _cell_style(amt, align="right", bg=bg)
        amt.number_format = '#,##0.00'

    return start_row + len(expense_rows) + 2


# ─── Public Function ──────────────────────────────────────────

def generate_report(telegram_id: int, date_from: str, date_to: str) -> io.BytesIO:
    summary_rows  = get_category_summary(telegram_id, date_from, date_to)
    expense_rows  = get_expenses_by_period(telegram_id, date_from, date_to)
    income_rows   = get_income_by_period(telegram_id, date_from, date_to)
    total_income  = get_income_summary(telegram_id, date_from, date_to)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Expense Report"

    # Banner
    sheet.merge_cells("A1:D1")
    banner = sheet["A1"]
    banner.value = f"Financial Report  |  {date_from}  to  {date_to}"
    banner.font = Font(bold=True, name="Arial", size=14, color="FFFFFF")
    banner.fill = PatternFill("solid", start_color="1A1A2E")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.append([])

    next_row = _build_summary_table(sheet, summary_rows, total_income, start_row=3)
    next_row = _build_income_table(sheet, income_rows, start_row=next_row)
    _build_expenses_table(sheet, expense_rows, start_row=next_row)

    _set_col_width(sheet, 1, 16)
    _set_col_width(sheet, 2, 35)
    _set_col_width(sheet, 3, 18)
    _set_col_width(sheet, 4, 18)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer